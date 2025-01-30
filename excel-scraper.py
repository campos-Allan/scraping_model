import os
import pandas as pd
from openpyxl import load_workbook
PATH = os.getcwd()+'\\excel_files\\'
FILES_EXCEL = [f for f in os.listdir(PATH) if 'xlsx' in f]
dados = {}
info = {}
CATEGORIA = ''
saverow = []

for file in FILES_EXCEL:
    workbook = load_workbook(filename=PATH+file, data_only=True)
    if 'template' not in file:
        sheet = workbook.active
        if 'dados' in file:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[-1] is None:
                    dados[row[3]] = [row[4], row[5], row[-2],
                                     row[-1], file.split(' ')[-1][:-5]]
            dados_df = pd.DataFrame.from_dict(
                dados).T.reset_index(drop=True)
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0] is None:
                    if isinstance(row[0], str):
                        CATEGORIA = row[0].split(' ')[-1].capitalize()
                    else:
                        info[row[0]] = [CATEGORIA, round(row[1] /
                                        row[2], 1), file.split(' ')[-1][:-5]]
            info_df = pd.DataFrame.from_dict(
                info).T.reset_index(drop=True)
    else:
        sheet_name = workbook.sheetnames
        for sheet_n in sheet_name:
            sheet = workbook[sheet_n]
            if sheet_n == 'DADOS':
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        linha = list(row[1:])
                        for i in range(0, len(linha)):
                            if linha[i] is None:
                                linha[i] = categ
                            else:
                                categ = linha[i]
                                linha[i] = linha[i]
                        saverow.append(linha)
                    else:
                        lista_condições = []
                        saverow[1] = ["." + item + '.' for item in saverow[1]]
                        for a, b, c in zip(saverow[0], saverow[1], saverow[2]):
                            lista_condições.append(a+b+c)
                        cond_df = pd.DataFrame(lista_condições)
                        for r_index, row in enumerate(sheet.iter_rows(min_row=5)):
                            if row[0].value is None:
                                break
                            cond_df['Valor'] = cond_df.map(lambda x: dados_df[(dados_df[3] == x.split('.')[1])
                                                                              & (dados_df[2] == x.split('.')[-1])
                                                                              & (dados_df[0] == x.split('.')[0])
                                                                              & (dados_df[4] == row[0].value.date().strftime('%d.%m.%y'))][1].sum())
                            for c_index, cell in enumerate(row):
                                if c_index > 0:
                                    sheet[cell.column_letter +
                                          str(cell.row)] = cond_df['Valor'][c_index-1]
                            cond_df = cond_df.drop('Valor', axis=1)
                            workbook.save(
                                filename="excel_files\\template.xlsx")
                        break
            else:
                info_df = info_df.groupby([0, 2])[1].sum().reset_index()
                for r_index, row in enumerate(sheet.iter_rows(min_row=3)):
                    if row[0].value is None:
                        break
                    insert = info_df[info_df[2] == row[0].value.date().strftime(
                        '%d.%m.%y')].reset_index(drop=True)[1]
                    if insert.empty:
                        break
                    for c_index, cell in enumerate(row):
                        if c_index > 0:
                            sheet[cell.column_letter +
                                  str(cell.row)] = insert[c_index-1]
                    workbook.save(
                        filename="excel_files\\template.xlsx")

        workbook.close()
