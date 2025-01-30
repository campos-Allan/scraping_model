import os
import pandas as pd
from tabula import read_pdf
from openpyxl import load_workbook
PATH = os.getcwd()+'\\pdf_files\\'
FILES_EXCEL = [f for f in os.listdir(PATH) if 'xlsx' in f]
FILES_PDF = [f for f in os.listdir(PATH) if 'pdf' in f]
dados_df = pd.DataFrame()
info_df = pd.DataFrame()
CATEGORIA = ''
saverow = []

for file in FILES_PDF:
    if 'dados' in file:
        content = pd.DataFrame(read_pdf(PATH+file, stream=True)[0])
        content = content[(content['OBS'] == 'Excedeu o Orçamento') | (
            content['OBS'] == 'Problemas na Conclusão')]
        try:
            content = content[['Categoria',
                               'Unnamed: 2', 'Responsável', 'OBS']]
            content = content.rename(
                columns={'Categoria': 0, 'Unnamed: 2': 1, 'Responsável': 2, 'OBS': 3})
        except KeyError:
            content = content[['Categoria',
                               'Unnamed: 1', 'Responsável', 'OBS']]
            content = content.rename(
                columns={'Categoria': 0, 'Unnamed: 1': 1, 'Responsável': 2, 'OBS': 3})
        content[4] = file.split(' ')[-1][:-4]
        content[1] = content[1].map(lambda x: ''.join(
            x.split(' ')).replace('.', '').replace(',00', '')).astype(int)
        content = content.reset_index(drop=True)
        dados_df = pd.concat([dados_df, content])
for file in FILES_EXCEL:
    workbook = load_workbook(filename=PATH+file, data_only=True)
    sheet_name = workbook.sheetnames
    for sheet_n in sheet_name:
        sheet = workbook[sheet_n]
        if sheet_n == 'DADOS':
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    linha = list(row[1:])
                    for i in range(0, len(linha)):
                        if linha[i] is None:
                            linha[i] = categ
                        else:
                            categ = linha[i]
                            linha[i] = linha[i]
                    saverow.append(linha)
                else:
                    lista_condições = []
                    saverow[1] = ["." + item + '.' for item in saverow[1]]
                    for a, b, c in zip(saverow[0], saverow[1], saverow[2]):
                        lista_condições.append(a+b+c)
                    cond_df = pd.DataFrame(lista_condições)
                    for r_index, row in enumerate(sheet.iter_rows(min_row=5)):
                        if row[0].value is None:
                            break
                        cond_df['Valor'] = cond_df.map(lambda x: dados_df[(dados_df[3] == x.split('.')[1])
                                                                          & (dados_df[2] == x.split('.')[-1])
                                                                          & (dados_df[0] == x.split('.')[0])
                                                                          & (dados_df[4] == row[0].value.date().strftime('%d.%m.%y'))][1].sum())
                        for c_index, cell in enumerate(row):
                            if c_index > 0:
                                sheet[cell.column_letter +
                                      str(cell.row)] = cond_df['Valor'][c_index-1]
                        cond_df = cond_df.drop('Valor', axis=1)
                        workbook.save(
                            filename="pdf_files\\template.xlsx")
                    break

    workbook.close()
